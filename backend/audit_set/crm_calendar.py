"""Auditor calendar projection and readable Excel export helpers."""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timezone
from typing import Any, Iterable


STANDARD_LABELS = {
    "QMS": "ISO 9001:2015",
    "EMS": "ISO 14001:2015",
    "OHSMS": "ISO 45001:2018",
    "FSMS": "ISO 22000:2018",
    "ISMS": "ISO/IEC 27001:2022",
    "MDQMS": "ISO 13485:2016",
    "ABMS": "ISO 37001:2016",
    "ENMS": "ISO 50001:2018",
}


def _value(obj: Any, key: str, default: Any = None) -> Any:
    return obj.get(key, default) if isinstance(obj, dict) else getattr(obj, key, default)


def _member_matches(member: Any, auditor_id: str) -> bool:
    return str(_value(member, "id", "") or "") == str(auditor_id)


def assignment_role(stage: Any, auditor_id: str) -> str | None:
    """Return the auditor's highest-priority assignment on a stage."""
    if str(_value(stage, "lead_auditor_id", "") or "") == str(auditor_id):
        return "Lead Auditor"
    role_groups = (
        ("auditors", "Auditor"),
        ("technical_experts", "Technical Expert"),
        ("trainees", "Trainee Auditor"),
        ("observers", "Observer"),
    )
    for field, label in role_groups:
        members = _value(stage, field, []) or []
        if any(_member_matches(member, auditor_id) for member in members):
            return label
    return None


def stage_label(value: str | None) -> str:
    raw = (value or "").strip()
    normalized = raw.lower().replace("-", "_").replace(" ", "_")
    if normalized in {"stage_1", "stage1", "1"}:
        return "Stage 1"
    if normalized in {"stage_2", "stage2", "2"}:
        return "Stage 2"
    if "surveillance" in normalized:
        if re.search(r"(?:^|_)1(?:_|$)", normalized):
            return "Surveillance 1"
        if re.search(r"(?:^|_)2(?:_|$)", normalized):
            return "Surveillance 2"
        return "Surveillance"
    if "recert" in normalized:
        return "Recertification"
    return raw or "Audit"


def audit_type_label(value: str | None) -> str:
    normalized = (value or "").strip().lower().replace("-", "_").replace(" ", "_")
    if normalized in {"initial", "initial_certification", "certification"}:
        return "Initial Certification"
    if "surveillance" in normalized:
        return "Surveillance"
    if "recert" in normalized:
        return "Recertification"
    if "transfer" in normalized:
        return "Transfer Certification"
    return (value or "Audit").replace("_", " ").title()


def standard_labels(values: Iterable[Any] | None) -> list[str]:
    labels: list[str] = []
    for value in values or []:
        raw = str(value or "").strip()
        if not raw:
            continue
        label = STANDARD_LABELS.get(raw.upper(), raw)
        if label not in labels:
            labels.append(label)
    return labels


def location_label(audit_set: Any) -> str:
    locations: list[str] = []
    for site in _value(audit_set, "sites", []) or []:
        address = str(_value(site, "address", "") or "").strip()
        name = str(_value(site, "name", "") or "").strip()
        label = " — ".join(part for part in (name, address) if part)
        if label and label not in locations:
            locations.append(label)
    if locations:
        return "; ".join(locations)
    address = str(_value(audit_set, "company_address", "") or "").strip()
    city = str(_value(audit_set, "city", "") or "").strip()
    country = str(_value(audit_set, "country", "") or "").strip()
    if address:
        return address
    return ", ".join(part for part in (city, country) if part)


def calendar_rows(stages: Iterable[Any], auditor_id: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for stage in stages:
        role = assignment_role(stage, auditor_id)
        if not role:
            continue
        audit_set = _value(stage, "audit_set")
        start = _value(stage, "audit_date_start")
        if not audit_set or not start:
            continue
        end = _value(stage, "audit_date_end") or start
        if end < start:
            start, end = end, start
        audit_type = (
            "Transfer Certification"
            if bool(_value(audit_set, "is_transfer", False))
            else audit_type_label(_value(audit_set, "audit_type"))
        )
        rows.append({
            "audit_set_id": str(_value(audit_set, "id", "")),
            "plan_number": int(_value(audit_set, "plan_number", 0) or 0),
            "company_name": str(_value(audit_set, "company_name", "") or ""),
            "location": location_label(audit_set),
            "standards": standard_labels(_value(audit_set, "standards", [])),
            "audit_type": audit_type,
            "stage_type": stage_label(_value(stage, "stage_type")),
            "date_start": start,
            "date_end": end,
            "audit_days": (end - start).days + 1,
            "auditor_role": role,
        })
    rows.sort(key=lambda row: (row["date_start"], row["company_name"], row["stage_type"]))
    return rows


def readable_date_range(start: date, end: date) -> str:
    if start == end:
        return f"{start.day} {start.strftime('%B %Y')}"
    if start.year == end.year and start.month == end.month:
        return f"{start.day}–{end.day} {start.strftime('%B %Y')}"
    if start.year == end.year:
        return f"{start.day} {start.strftime('%B')}–{end.day} {end.strftime('%B %Y')}"
    return f"{start.day} {start.strftime('%B %Y')}–{end.day} {end.strftime('%B %Y')}"


def build_auditor_schedule_workbook(auditor_name: str, rows: list[dict[str, Any]]) -> bytes:
    """Build a polished, occupied-days-only Excel schedule."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Audit Schedule"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"

    columns = [
        "Audit Dates", "Days", "Company / Client", "Location / Site", "Standards",
        "Audit Type", "Stage", "Assignment Role", "Plan Number",
        "Start Date (sort)", "End Date (sort)",
    ]
    last_column = "K"
    visible_last_column = "I"
    sheet.merge_cells(f"A1:{visible_last_column}1")
    sheet["A1"] = f"Auditor Schedule — {auditor_name}"
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1A4731")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells(f"A2:{visible_last_column}2")
    sheet["A2"] = (
        f"Occupied audit periods only · {len(rows)} assignment(s) · "
        f"Exported {datetime.now(timezone.utc).strftime('%d %B %Y %H:%M UTC')}"
    )
    sheet["A2"].font = Font(name="Aptos", size=10, italic=True, color="52645B")
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 22

    header_row = 4
    for column, header in enumerate(columns, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="25634A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 30

    thin_green = Side(style="thin", color="C9D9D0")
    for row_index, row in enumerate(rows, start=5):
        values = [
            readable_date_range(row["date_start"], row["date_end"]),
            row["audit_days"],
            row["company_name"],
            row["location"],
            " + ".join(row["standards"]),
            row["audit_type"],
            row["stage_type"],
            row["auditor_role"],
            row["plan_number"],
            row["date_start"],
            row["date_end"],
        ]
        for column, value in enumerate(values, start=1):
            if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                value = "'" + value
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="Aptos", size=10, color="1F2937")
            cell.alignment = Alignment(vertical="top", wrap_text=column in {1, 3, 4, 5, 6, 7, 8})
            cell.border = Border(bottom=thin_green)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F7F4")
        sheet.cell(row_index, 2).number_format = "0"
        sheet.cell(row_index, 9).number_format = '"#"0'
        sheet.cell(row_index, 10).number_format = "dd mmm yyyy"
        sheet.cell(row_index, 11).number_format = "dd mmm yyyy"
        sheet.row_dimensions[row_index].height = 34

    if rows:
        table = Table(displayName="AuditorSchedule", ref=f"A4:{last_column}{4 + len(rows)}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.merge_cells(f"A5:{visible_last_column}5")
        sheet["A5"] = "No scheduled audit assignments were found for this auditor."
        sheet["A5"].font = Font(name="Aptos", italic=True, color="6B7280")
        sheet["A5"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[5].height = 32

    widths = {
        "A": 24, "B": 8, "C": 28, "D": 38, "E": 34,
        "F": 23, "G": 20, "H": 20, "I": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions["J"].hidden = True
    sheet.column_dimensions["K"].hidden = True

    sheet.auto_filter.ref = f"A4:{last_column}{max(4, 4 + len(rows))}"
    sheet.print_title_rows = "1:4"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = f"A1:{visible_last_column}{max(5, 4 + len(rows))}"
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.oddFooter.center.text = "Certiva · Auditor Schedule"
    sheet.oddFooter.right.text = "Page &P of &N"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def safe_export_name(auditor_name: str) -> str:
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", auditor_name.strip()).strip("_")
    return f"Auditor_Schedule_{safe_name or 'Auditor'}_{date.today().strftime('%Y%m%d')}.xlsx"
