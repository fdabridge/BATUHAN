"""
CRM Router — Portal 91

Read-only endpoints for non-technical staff (finance / operations).
Data is derived from the existing audit_set tables with no writes.

Accessible to roles: crm, admin
All endpoints return empty/zero responses on DB error — never 500.
"""
from __future__ import annotations

import logging
from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session

from audit_set.db_models import AuditSet, AuditSetStage, CRMCertificateCommercial, get_db
from auth.db_models import PlatformUser, get_db as get_auth_db
from auth.dependencies import get_current_user
from auditors.models import Auditor, get_db as get_auditors_db

logger = logging.getLogger(__name__)
router = APIRouter(tags=["crm"])

CRM_ROLES = {"crm", "admin"}

# ── Simplified pipeline status map ───────────────────────────────────────────

SIMPLE_STATUS: dict[str | None, str] = {
    None:                    "Application Received",
    "pending_review":        "Application Received",
    "in_planning":           "In Planning",
    "notification_sent":     "In Planning",
    "quotation_sent":        "Quotation Sent",
    "agreement_signed":      "Agreement Signed",
    "fr218_in_progress":     "Under Review",
    "fr218_complete":        "Under Review",
    "stage1_scheduled":      "Stage 1 Audit",
    "stage1_in_progress":    "Stage 1 Audit",
    "stage1_complete":       "Stage 1 Complete",
    "stage2_scheduled":      "Stage 2 Audit",
    "stage2_in_progress":    "Stage 2 Audit",
    "under_review":          "Under Review",
    "committee_review":      "Committee Review",
    "audit_scheduled":       "Surveillance Audit",
    "audit_in_progress":     "Surveillance Audit",
    "surveillance_complete": "Surveillance Complete",
    "cert_complete":         "Certified",
    "certified":             "Certified",
}

PIPELINE_ORDER = [
    "Application Received", "In Planning", "Quotation Sent", "Agreement Signed",
    "Under Review", "Stage 1 Audit", "Stage 1 Complete", "Stage 2 Audit",
    "Committee Review", "Surveillance Audit", "Surveillance Complete", "Certified",
]

# ── Cycle date computation ────────────────────────────────────────────────────

def _add_years(d: date, years: int) -> date:
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        return d.replace(year=d.year + years, day=28)


def _cycle_dates(issued: date) -> tuple[date, date, date]:
    surv1  = _add_years(issued, 1) - timedelta(days=1)
    surv2  = _add_years(issued, 2) - timedelta(days=1)
    recert = _add_years(issued, 3)
    return surv1, surv2, recert


# ── Contact lookup ────────────────────────────────────────────────────────────

def _get_contact(audit_set_id: str, audit_set: AuditSet, auth_db: Session) -> dict:
    try:
        client_user = (
            auth_db.query(PlatformUser)
            .filter_by(audit_set_id=audit_set_id, role="client")
            .first()
        )
        if client_user:
            return {
                "contact_name":  client_user.full_name,
                "contact_email": client_user.email,
                "contact_phone": audit_set.phone or "",
            }
    except Exception:
        pass
    return {
        "contact_name":  audit_set.representative or "",
        "contact_email": audit_set.email or "",
        "contact_phone": audit_set.phone or "",
    }


# ── Serialise one AuditSet row ────────────────────────────────────────────────

def _serialise(audit_set: AuditSet, auth_db: Session) -> dict:
    issued: Optional[date] = audit_set.cert_issued_date
    surv1_due = surv2_due = recert_due = None
    if issued:
        surv1_due, surv2_due, recert_due = _cycle_dates(issued)
    contact = _get_contact(audit_set.id, audit_set, auth_db)

    # Consultant lookup
    consultant_name: str | None = None
    if audit_set.consultant_id:
        try:
            c = auth_db.query(PlatformUser).filter_by(id=audit_set.consultant_id).first()
            if c:
                consultant_name = c.full_name
        except Exception:
            pass

    return {
        "id":                audit_set.id,
        "company_name":      audit_set.company_name or "",
        "city":              audit_set.city or "",
        "standards":         audit_set.standards or [],
        "audit_type":        audit_set.audit_type or "initial",
        "accreditation_body": audit_set.accreditation_body or "",
        "simple_status":     SIMPLE_STATUS.get(audit_set.workflow_status, "In Progress"),
        "workflow_status":   audit_set.workflow_status,
        "cert_issued_date":  issued.isoformat() if issued else None,
        "cert_expiry_date":  audit_set.cert_expiry_date.isoformat() if audit_set.cert_expiry_date else None,
        "surv1_due":         surv1_due.isoformat()  if surv1_due  else None,
        "surv2_due":         surv2_due.isoformat()  if surv2_due  else None,
        "recert_due":        recert_due.isoformat() if recert_due else None,
        "certification_fee": audit_set.certification_fee,
        "surveillance_fee":  audit_set.surveillance_fee,
        "currency":          audit_set.currency or "USD",
        "consultant_id":     audit_set.consultant_id,
        "consultant_name":   consultant_name or ("IFC Global" if not audit_set.consultant_id else None),
        **contact,
    }


# ── Pydantic response schemas ─────────────────────────────────────────────────

class CRMClientRow(BaseModel):
    id: str
    company_name: str
    city: str
    standards: list
    audit_type: str
    accreditation_body: str
    simple_status: str
    workflow_status: Optional[str]
    cert_issued_date: Optional[str]
    cert_expiry_date: Optional[str]
    surv1_due: Optional[str]
    surv2_due: Optional[str]
    recert_due: Optional[str]
    certification_fee: Optional[float]
    surveillance_fee: Optional[float]
    currency: str
    contact_name: str
    contact_email: str
    contact_phone: str
    consultant_id:   Optional[str]
    consultant_name: Optional[str]

    class Config:
        from_attributes = True


class ConsultantAssignmentRequest(BaseModel):
    consultant_id: Optional[str] = None


class KPIs(BaseModel):
    active_certifications: int
    expiring_90_days: int
    overdue_renewals: int
    in_progress: int


class CRMDashboardResponse(BaseModel):
    kpis: KPIs
    upcoming_renewals: list[dict]
    pipeline: dict[str, int]


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/crm/dashboard", response_model=CRMDashboardResponse)
def crm_dashboard(
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    if current_user.role not in CRM_ROLES:
        from fastapi import HTTPException
        raise HTTPException(403, "Not authorized")

    today = date.today()
    in_90 = today + timedelta(days=90)

    try:
        all_sets = db.query(AuditSet).all()
    except Exception as exc:
        logger.error("[CRM] dashboard DB error: %s", exc)
        all_sets = []

    active_certs = expiring_90 = overdue_renewals = in_progress_cnt = 0
    pipeline: dict[str, int] = {s: 0 for s in PIPELINE_ORDER}
    upcoming: list[dict] = []

    for a in all_sets:
        label = SIMPLE_STATUS.get(a.workflow_status, "In Progress")
        if label in pipeline:
            pipeline[label] += 1
        if label not in ("Certified", "Application Received"):
            in_progress_cnt += 1
        if not a.cert_issued_date:
            continue

        active_certs += 1
        issued = a.cert_issued_date
        surv1, surv2, recert = _cycle_dates(issued)
        contact = _get_contact(a.id, a, auth_db)

        if a.cert_expiry_date and today <= a.cert_expiry_date <= in_90:
            expiring_90 += 1

        for milestone, due in [("surv1", surv1), ("surv2", surv2), ("recert", recert)]:
            days_until = (due - today).days
            if days_until < -30:
                overdue_renewals += 1
                continue
            if days_until > 548:
                continue
            upcoming.append({
                "audit_set_id":     a.id,
                "company_name":     a.company_name or "",
                "standards":        a.standards or [],
                "milestone":        milestone,
                "due_date":         due.isoformat(),
                "days_until":       days_until,
                "cert_issued_date": issued.isoformat(),
                **contact,
            })

    upcoming.sort(key=lambda r: r["due_date"])

    return CRMDashboardResponse(
        kpis=KPIs(
            active_certifications=active_certs,
            expiring_90_days=expiring_90,
            overdue_renewals=overdue_renewals,
            in_progress=in_progress_cnt,
        ),
        upcoming_renewals=upcoming,
        pipeline={k: v for k, v in pipeline.items() if v > 0},
    )


@router.get("/crm/clients", response_model=list[CRMClientRow])
def crm_clients(
    consultant_id: Optional[str] = None,   # filter by consultant; "none" = IFC Global direct
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    if current_user.role not in CRM_ROLES:
        from fastapi import HTTPException
        raise HTTPException(403, "Not authorized")

    try:
        q = db.query(AuditSet).order_by(AuditSet.company_name)
        if consultant_id == "none":
            q = q.filter(AuditSet.consultant_id.is_(None))
        elif consultant_id:
            q = q.filter(AuditSet.consultant_id == consultant_id)
        all_sets = q.all()
    except Exception as exc:
        logger.error("[CRM] clients DB error: %s", exc)
        return []

    result = []
    for a in all_sets:
        try:
            result.append(CRMClientRow(**_serialise(a, auth_db)))
        except Exception as exc:
            logger.warning("[CRM] skip audit_set %s: %s", a.id, exc)
    return result


@router.get("/crm/clients/{audit_set_id}", response_model=CRMClientRow)
def crm_client_detail(
    audit_set_id: str,
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    if current_user.role not in CRM_ROLES:
        from fastapi import HTTPException
        raise HTTPException(403, "Not authorized")

    try:
        a = db.query(AuditSet).filter_by(id=audit_set_id).first()
    except Exception as exc:
        logger.error("[CRM] client detail DB error: %s", exc)
        from fastapi import HTTPException
        raise HTTPException(500, "Database error")

    if not a:
        from fastapi import HTTPException
        raise HTTPException(404, "Audit set not found")

    return CRMClientRow(**_serialise(a, auth_db))


@router.patch("/crm/clients/{audit_set_id}/consultant", response_model=CRMClientRow)
def update_client_consultant(
    audit_set_id: str,
    body: ConsultantAssignmentRequest,
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    """Admin-only: attach, change, or clear a consultant for an existing client."""
    if current_user.role != "admin":
        raise HTTPException(403, "Only admins can change client consultant assignments")

    audit_set = db.query(AuditSet).filter_by(id=audit_set_id).first()
    if not audit_set:
        raise HTTPException(404, "Audit set not found")

    consultant_id = (body.consultant_id or "").strip() or None
    if consultant_id:
        consultant = (
            auth_db.query(PlatformUser)
            .filter(
                PlatformUser.id == consultant_id,
                PlatformUser.role == "consultant",
                PlatformUser.is_active == True,
            )
            .first()
        )
        if not consultant:
            raise HTTPException(400, "Selected consultant is not an active consultant user")

    audit_set.consultant_id = consultant_id
    db.commit()
    db.refresh(audit_set)
    return CRMClientRow(**_serialise(audit_set, auth_db))


# ── Portal 92 — Auditor Calendar ──────────────────────────────────────────────

class CRMAuditorRow(BaseModel):
    id:    str
    name:  str
    email: Optional[str]
    role:  Optional[str]


class CRMCalendarEntry(BaseModel):
    audit_set_id:  str
    plan_number:   int
    company_name:  str
    stage_type:    str
    date_start:    str
    date_end:      str
    auditor_role:  str


@router.get("/crm/auditors", response_model=list[CRMAuditorRow])
def crm_auditors(
    auditors_db: Session = Depends(get_auditors_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    if current_user.role not in CRM_ROLES:
        from fastapi import HTTPException
        raise HTTPException(403, "Not authorized")
    try:
        rows = (
            auditors_db.query(Auditor)
            .filter(Auditor.is_active == True)
            .order_by(Auditor.name)
            .all()
        )
        return [CRMAuditorRow(id=r.id, name=r.name, email=r.email, role=r.role) for r in rows]
    except Exception as exc:
        logger.error("[CRM] auditors DB error: %s", exc)
        return []


@router.get("/crm/auditors/{auditor_id}/calendar", response_model=list[CRMCalendarEntry])
def crm_auditor_calendar(
    auditor_id: str,
    db: Session = Depends(get_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    if current_user.role not in CRM_ROLES:
        from fastapi import HTTPException
        raise HTTPException(403, "Not authorized")
    try:
        stages = (
            db.query(AuditSetStage)
            .join(AuditSet, AuditSetStage.audit_set_id == AuditSet.id)
            .filter(AuditSetStage.audit_date_start.isnot(None))
            .all()
        )
    except Exception as exc:
        logger.error("[CRM] calendar DB error: %s", exc)
        return []

    result: list[CRMCalendarEntry] = []
    for stage in stages:
        is_lead = stage.lead_auditor_id == auditor_id
        team: list[dict] = stage.auditors or []
        is_team = any(str(m.get("id", "")) == auditor_id for m in team if isinstance(m, dict))
        if not (is_lead or is_team):
            continue

        audit_set = stage.audit_set
        if not audit_set:
            continue

        date_start = stage.audit_date_start
        date_end   = stage.audit_date_end or date_start

        stype = (stage.stage_type or "").lower()
        if "stage_1" in stype or stype in ("stage1", "1"):
            label = "Stage 1"
        elif "stage_2" in stype or stype in ("stage2", "2"):
            label = "Stage 2"
        elif "surveillance" in stype:
            label = "Surveillance"
        elif "recert" in stype:
            label = "Recertification"
        else:
            label = stage.stage_type or "Audit"

        result.append(CRMCalendarEntry(
            audit_set_id = audit_set.id,
            plan_number  = audit_set.plan_number,
            company_name = audit_set.company_name or "",
            stage_type   = label,
            date_start   = date_start.isoformat(),
            date_end     = date_end.isoformat(),
            auditor_role = "Lead Auditor" if is_lead else "Team Auditor",
        ))

    result.sort(key=lambda r: r.date_start)
    return result


# ── Portal 106 — Consultant revenue summary ─────────────────────────────────

class ConsultantSummary(BaseModel):
    id:               str
    full_name:        str
    username:         Optional[str]
    email:            str
    client_count:     int
    certified_count:  int
    total_revenue:    float          # sum of certification_fee + surveillance_fee
    renewals_90_days: int            # clients with any milestone due in next 90 days


@router.get("/crm/consultants", response_model=list[ConsultantSummary])
def crm_consultants(
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    """Revenue and client summary per consultant. CRM / admin only."""
    if current_user.role not in CRM_ROLES:
        from fastapi import HTTPException
        raise HTTPException(403, "Not authorized")

    today = date.today()
    in_90 = today + timedelta(days=90)

    try:
        consultants = (
            auth_db.query(PlatformUser)
            .filter_by(role="consultant", is_active=True)
            .order_by(PlatformUser.full_name)
            .all()
        )
    except Exception as exc:
        logger.error("[CRM] consultants DB error: %s", exc)
        return []

    result: list[ConsultantSummary] = []
    for c in consultants:
        try:
            audit_sets = db.query(AuditSet).filter_by(consultant_id=c.id).all()
            certified  = [a for a in audit_sets if a.workflow_status == "certified"]
            revenue    = sum(
                (a.certification_fee or 0) + (a.surveillance_fee or 0)
                for a in audit_sets
            )
            renewals = 0
            for a in audit_sets:
                if not a.cert_issued_date:
                    continue
                s1, s2, rc = _cycle_dates(a.cert_issued_date)
                if any(today <= d <= in_90 for d in [s1, s2, rc]):
                    renewals += 1
            result.append(ConsultantSummary(
                id=c.id,
                full_name=c.full_name,
                username=c.username,
                email=c.email,
                client_count=len(audit_sets),
                certified_count=len(certified),
                total_revenue=revenue,
                renewals_90_days=renewals,
            ))
        except Exception as exc:
            logger.warning("[CRM] consultant %s summary error: %s", c.id, exc)
    return result


# ── Portal 91b — CRM Certificate Cockpit Schemas ─────────────────────────────

class CertificateRow(BaseModel):
    audit_set_id: str
    plan_number: int
    company_name: str
    standard: str
    certificate_number: Optional[str]  # plan_number-standard e.g. "1652-QMS"
    lifecycle_status: str  # active, expiring_soon, expired, suspended, withdrawn, in_progress
    cert_issued_date: Optional[str]
    cert_expiry_date: Optional[str]
    next_surveillance_due: Optional[str]
    countdown_days: Optional[int]
    last_surveillance_completed: Optional[str]
    payment_status: str
    amount_due: Optional[float]
    amount_received: Optional[float]
    outstanding: Optional[float]
    notes: Optional[str]
    consultant_name: Optional[str]
    assigned_auditor: Optional[str]

    class Config:
        from_attributes = True


class CertificateDashboardSummary(BaseModel):
    active_certificates: int
    expiring_in_90_days: int
    surveillance_due_in_30_days: int
    overdue_surveillance: int
    expired: int
    total_outstanding: float
    total_collected: float


class CommercialUpdateRequest(BaseModel):
    standard: str
    payment_status: Optional[str] = None
    amount_due: Optional[float] = None
    amount_received: Optional[float] = None
    notes: Optional[str] = None


class CertificateCockpitResponse(BaseModel):
    summary: CertificateDashboardSummary
    certificates: list[CertificateRow]


# ── Portal 91b — CRM Certificate Cockpit ────────────────────────────────────

def _compute_lifecycle_status(audit_set: AuditSet) -> str:
    """Derive lifecycle status from audit set data. CRM cannot edit this."""
    if audit_set.cert_status == "suspended":
        return "suspended"
    if audit_set.cert_status == "withdrawn":
        return "withdrawn"
    if not audit_set.cert_issued_date:
        return "in_progress"
    today = date.today()
    if audit_set.cert_expiry_date and audit_set.cert_expiry_date < today:
        return "expired"
    if audit_set.cert_expiry_date and (audit_set.cert_expiry_date - today).days <= 90:
        return "expiring_soon"
    return "active"


def _compute_surveillance_due(audit_set: AuditSet, db: Session) -> tuple[Optional[date], Optional[date]]:
    """
    Compute next surveillance due date based on actual audit stage completions.
    Rules:
    - First surveillance: within 12 months of Stage 2 end date
    - Later surveillances: within 12 months of previous completed surveillance end date
    Returns: (next_surveillance_due, last_surveillance_completed)
    """
    stages = (
        db.query(AuditSetStage)
        .filter_by(audit_set_id=audit_set.id)
        .order_by(AuditSetStage.stage_order)
        .all()
    )

    # Find completed surveillance stages (most recent first)
    completed_surveillances = [
        s for s in stages
        if s.status == "complete"
        and s.stage_type and "surveillance" in s.stage_type.lower()
        and s.audit_date_end
    ]
    completed_surveillances.sort(key=lambda s: s.audit_date_end, reverse=True)

    last_surv_completed: Optional[date] = None
    if completed_surveillances:
        last_surv_completed = completed_surveillances[0].audit_date_end

    # Find Stage 2 completion date
    stage2_end: Optional[date] = None
    for s in stages:
        if s.stage_type and ("stage_2" in s.stage_type.lower() or s.stage_type.lower() in ("stage2", "2")):
            if s.status == "complete" and s.audit_date_end:
                stage2_end = s.audit_date_end
                break

    # Compute next due
    next_due: Optional[date] = None
    if last_surv_completed:
        # Next surveillance due within 12 months of last completed surveillance
        next_due = _add_years(last_surv_completed, 1)
    elif stage2_end:
        # First surveillance due within 12 months of Stage 2 end
        next_due = _add_years(stage2_end, 1)

    return next_due, last_surv_completed


def _build_certificate_row(
    audit_set: AuditSet, standard: str, db: Session, auth_db: Session, commercial: Optional[CRMCertificateCommercial]
) -> CertificateRow:
    lifecycle = _compute_lifecycle_status(audit_set)
    next_surv_due, last_surv_completed = _compute_surveillance_due(audit_set, db)

    countdown: Optional[int] = None
    if next_surv_due:
        countdown = (next_surv_due - date.today()).days

    # Consultant name
    consultant_name: Optional[str] = None
    if audit_set.consultant_id:
        try:
            c = auth_db.query(PlatformUser).filter_by(id=audit_set.consultant_id).first()
            if c:
                consultant_name = c.full_name
        except Exception:
            pass

    # Assigned auditor (latest stage lead auditor)
    assigned_auditor: Optional[str] = None
    if audit_set.stages:
        for stage in reversed(audit_set.stages):
            if stage.lead_auditor_name:
                assigned_auditor = stage.lead_auditor_name
                break

    # Payment/commercial info
    payment_status = "unpaid"
    amount_due: Optional[float] = None
    amount_received: Optional[float] = None
    notes: Optional[str] = None
    if commercial:
        payment_status = commercial.payment_status
        amount_due = commercial.amount_due
        amount_received = commercial.amount_received
        notes = commercial.notes

    outstanding = None
    if amount_due is not None:
        outstanding = (amount_due or 0) - (amount_received or 0)

    return CertificateRow(
        audit_set_id=audit_set.id,
        plan_number=audit_set.plan_number,
        company_name=audit_set.company_name or "",
        standard=standard,
        certificate_number=f"{audit_set.plan_number}-{standard}" if audit_set.cert_issued_date else None,
        lifecycle_status=lifecycle,
        cert_issued_date=audit_set.cert_issued_date.isoformat() if audit_set.cert_issued_date else None,
        cert_expiry_date=audit_set.cert_expiry_date.isoformat() if audit_set.cert_expiry_date else None,
        next_surveillance_due=next_surv_due.isoformat() if next_surv_due else None,
        countdown_days=countdown,
        last_surveillance_completed=last_surv_completed.isoformat() if last_surv_completed else None,
        payment_status=payment_status,
        amount_due=amount_due,
        amount_received=amount_received,
        outstanding=outstanding,
        notes=notes,
        consultant_name=consultant_name,
        assigned_auditor=assigned_auditor,
    )


@router.get("/crm/certificates", response_model=CertificateCockpitResponse)
def crm_certificates(
    status: Optional[str] = None,        # lifecycle filter
    standard: Optional[str] = None,      # e.g. "QMS"
    payment: Optional[str] = None,       # payment_status filter
    consultant_id: Optional[str] = None,
    overdue_bucket: Optional[str] = None,  # "due_this_month" | "due_30" | "overdue" | "recert_due" | "expiring_soon" | "expired"
    db: Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    """One row per certificate/standard. Portfolio view for CRM cockpit."""
    if current_user.role not in CRM_ROLES:
        raise HTTPException(403, "Not authorized")

    try:
        q = db.query(AuditSet).order_by(AuditSet.company_name)
        if consultant_id:
            if consultant_id == "none":
                q = q.filter(AuditSet.consultant_id.is_(None))
            else:
                q = q.filter(AuditSet.consultant_id == consultant_id)
        all_sets = q.all()
    except Exception as exc:
        logger.error("[CRM] certificates DB error: %s", exc)
        all_sets = []

    # Load all commercial records
    try:
        commercials = db.query(CRMCertificateCommercial).all()
        commercial_map: dict[str, CRMCertificateCommercial] = {
            f"{c.audit_set_id}:{c.standard}": c for c in commercials
        }
    except Exception:
        commercial_map = {}

    today = date.today()
    certificates: list[CertificateRow] = []

    # Summary accumulators
    active_count = expiring_90 = surv_due_30 = overdue_surv = expired_count = 0
    total_outstanding = 0.0
    total_collected = 0.0

    for a in all_sets:
        standards_list = a.standards or []
        if not standards_list:
            standards_list = ["N/A"]

        for std in standards_list:
            if standard and std.upper() != standard.upper():
                continue

            key = f"{a.id}:{std}"
            commercial = commercial_map.get(key)

            row = _build_certificate_row(a, std, db, auth_db, commercial)

            # Apply filters
            if status and row.lifecycle_status != status:
                continue
            if payment and row.payment_status != payment:
                continue

            # Overdue bucket filter
            if overdue_bucket:
                if overdue_bucket == "due_this_month":
                    if not row.next_surveillance_due:
                        continue
                    due_d = date.fromisoformat(row.next_surveillance_due)
                    if not (due_d.year == today.year and due_d.month == today.month):
                        continue
                elif overdue_bucket == "due_30":
                    if row.countdown_days is None or row.countdown_days < 0 or row.countdown_days > 30:
                        continue
                elif overdue_bucket == "overdue":
                    if row.countdown_days is None or row.countdown_days >= 0:
                        continue
                elif overdue_bucket == "recert_due":
                    if row.lifecycle_status not in ("expiring_soon", "expired"):
                        continue
                elif overdue_bucket == "expiring_soon":
                    if row.lifecycle_status != "expiring_soon":
                        continue
                elif overdue_bucket == "expired":
                    if row.lifecycle_status != "expired":
                        continue

            certificates.append(row)

            # Summary calculations
            if row.lifecycle_status == "active":
                active_count += 1
            elif row.lifecycle_status == "expiring_soon":
                expiring_90 += 1
            elif row.lifecycle_status == "expired":
                expired_count += 1

            if row.countdown_days is not None:
                if 0 <= row.countdown_days <= 30:
                    surv_due_30 += 1
                elif row.countdown_days < 0:
                    overdue_surv += 1

            if row.outstanding and row.outstanding > 0:
                total_outstanding += row.outstanding
            if row.amount_received:
                total_collected += row.amount_received

    summary = CertificateDashboardSummary(
        active_certificates=active_count,
        expiring_in_90_days=expiring_90,
        surveillance_due_in_30_days=surv_due_30,
        overdue_surveillance=overdue_surv,
        expired=expired_count,
        total_outstanding=round(total_outstanding, 2),
        total_collected=round(total_collected, 2),
    )

    return CertificateCockpitResponse(summary=summary, certificates=certificates)


@router.get("/crm/certificates/export")
def crm_certificates_export(
    db: Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    """
    Export Excel with three sheets: Portfolio, Radar (surveillance due), Financials.
    """
    if current_user.role not in CRM_ROLES:
        raise HTTPException(403, "Not authorized")

    import io
    from datetime import datetime as dt

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl not available
        raise HTTPException(500, "openpyxl package required for Excel export")

    # Build all certificate rows
    all_sets = db.query(AuditSet).order_by(AuditSet.company_name).all()
    commercials = db.query(CRMCertificateCommercial).all()
    commercial_map = {f"{c.audit_set_id}:{c.standard}": c for c in commercials}

    rows: list[CertificateRow] = []
    for a in all_sets:
        for std in (a.standards or ["N/A"]):
            key = f"{a.id}:{std}"
            commercial = commercial_map.get(key)
            rows.append(_build_certificate_row(a, std, db, auth_db, commercial))

    wb = openpyxl.Workbook()

    # Sheet 1 — Portfolio
    ws1 = wb.active
    ws1.title = "Portfolio"
    headers1 = ["Company", "Standard", "Certificate #", "Status", "Issued", "Expires",
                "Next Surv Due", "Countdown (days)", "Last Surv Completed", "Auditor", "Consultant"]
    ws1.append(headers1)
    for h_cell in ws1[1]:
        h_cell.font = Font(bold=True)
    for r in rows:
        ws1.append([
            r.company_name, r.standard, r.certificate_number or "", r.lifecycle_status,
            r.cert_issued_date or "", r.cert_expiry_date or "",
            r.next_surveillance_due or "", r.countdown_days if r.countdown_days is not None else "",
            r.last_surveillance_completed or "", r.assigned_auditor or "", r.consultant_name or ""
        ])

    # Sheet 2 — Radar (surveillance/renewals due)
    ws2 = wb.create_sheet("Radar")
    headers2 = ["Company", "Standard", "Certificate #", "Next Surv Due", "Countdown (days)", "Status", "Auditor"]
    ws2.append(headers2)
    for h_cell in ws2[1]:
        h_cell.font = Font(bold=True)
    radar_rows = [r for r in rows if r.countdown_days is not None and r.countdown_days <= 90]
    radar_rows.sort(key=lambda x: x.countdown_days if x.countdown_days is not None else 999)
    for r in radar_rows:
        ws2.append([
            r.company_name, r.standard, r.certificate_number or "",
            r.next_surveillance_due or "", r.countdown_days,
            r.lifecycle_status, r.assigned_auditor or ""
        ])

    # Sheet 3 — Financials
    ws3 = wb.create_sheet("Financials")
    headers3 = ["Company", "Standard", "Payment Status", "Amount Due", "Amount Received", "Outstanding", "Notes"]
    ws3.append(headers3)
    for h_cell in ws3[1]:
        h_cell.font = Font(bold=True)
    for r in rows:
        ws3.append([
            r.company_name, r.standard, r.payment_status,
            r.amount_due if r.amount_due is not None else "",
            r.amount_received if r.amount_received is not None else "",
            r.outstanding if r.outstanding is not None else "",
            r.notes or ""
        ])

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    filename = f"CRM_Certificates_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/crm/certificates/{audit_set_id}")
def crm_certificate_detail(
    audit_set_id: str,
    db: Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    """Returns certificate rows for a single audit set (one per standard)."""
    if current_user.role not in CRM_ROLES:
        raise HTTPException(403, "Not authorized")

    a = db.query(AuditSet).filter_by(id=audit_set_id).first()
    if not a:
        raise HTTPException(404, "Audit set not found")

    standards_list = a.standards or ["N/A"]
    commercials = (
        db.query(CRMCertificateCommercial)
        .filter_by(audit_set_id=audit_set_id)
        .all()
    )
    commercial_map = {c.standard: c for c in commercials}

    rows = []
    for std in standards_list:
        commercial = commercial_map.get(std)
        rows.append(_build_certificate_row(a, std, db, auth_db, commercial))

    return rows


@router.patch("/crm/certificates/{audit_set_id}/commercial")
def update_certificate_commercial(
    audit_set_id: str,
    body: CommercialUpdateRequest,
    db: Session = Depends(get_db),
    current_user: PlatformUser = Depends(get_current_user),
):
    """
    CRM users can ONLY update commercial fields: payment_status, amount_due, amount_received, notes.
    Cannot modify: lifecycle status, cert dates, audit history, standards, surveillance dates.
    """
    if current_user.role not in CRM_ROLES:
        raise HTTPException(403, "Not authorized")

    # Validate audit set exists
    audit_set = db.query(AuditSet).filter_by(id=audit_set_id).first()
    if not audit_set:
        raise HTTPException(404, "Audit set not found")

    # Validate standard is part of this audit set
    standards_list = audit_set.standards or []
    if body.standard not in standards_list:
        raise HTTPException(400, f"Standard '{body.standard}' not found in this audit set")

    # Validate payment_status enum
    valid_statuses = {"paid", "partially_paid", "unpaid", "overdue"}
    if body.payment_status and body.payment_status not in valid_statuses:
        raise HTTPException(400, f"Invalid payment_status. Must be one of: {', '.join(valid_statuses)}")

    # Upsert commercial record
    existing = (
        db.query(CRMCertificateCommercial)
        .filter_by(audit_set_id=audit_set_id, standard=body.standard)
        .first()
    )

    if existing:
        if body.payment_status is not None:
            existing.payment_status = body.payment_status
        if body.amount_due is not None:
            existing.amount_due = body.amount_due
        if body.amount_received is not None:
            existing.amount_received = body.amount_received
        if body.notes is not None:
            existing.notes = body.notes
        existing.updated_by = current_user.id
    else:
        existing = CRMCertificateCommercial(
            audit_set_id=audit_set_id,
            standard=body.standard,
            payment_status=body.payment_status or "unpaid",
            amount_due=body.amount_due,
            amount_received=body.amount_received,
            notes=body.notes,
            updated_by=current_user.id,
        )
        db.add(existing)

    db.commit()
    db.refresh(existing)

    return {
        "audit_set_id": existing.audit_set_id,
        "standard": existing.standard,
        "payment_status": existing.payment_status,
        "amount_due": existing.amount_due,
        "amount_received": existing.amount_received,
        "outstanding": (existing.amount_due or 0) - (existing.amount_received or 0),
        "notes": existing.notes,
        "updated_by": existing.updated_by,
    }
