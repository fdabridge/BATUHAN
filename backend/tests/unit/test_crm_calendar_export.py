"""Regression tests for occupied-days-only auditor calendar exports."""

from datetime import date
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from audit_set.crm_calendar import (
    assignment_role,
    build_auditor_schedule_workbook,
    calendar_rows,
    readable_date_range,
    stage_label,
)


def _audit_set(**overrides):
    values = {
        "id": "set-1",
        "plan_number": 1725,
        "company_name": "Example Foods",
        "company_address": "1 Factory Road, Istanbul",
        "city": "Istanbul",
        "country": "Türkiye",
        "sites": [{"name": "Main Factory", "address": "1 Factory Road, Istanbul"}],
        "standards": ["QMS", "FSMS"],
        "audit_type": "initial",
    }
    values.update(overrides)
    return SimpleNamespace(**values)


def _stage(**overrides):
    values = {
        "audit_set": _audit_set(),
        "audit_date_start": date(2026, 9, 1),
        "audit_date_end": date(2026, 9, 3),
        "stage_type": "stage_2",
        "lead_auditor_id": None,
        "auditors": [],
        "technical_experts": [],
        "trainees": [],
        "observers": [],
    }
    values.update(overrides)
    return SimpleNamespace(**values)


def test_assignment_roles_include_technical_experts_and_preserve_lead_priority():
    assert assignment_role(_stage(technical_experts=[{"id": "a1"}]), "a1") == "Technical Expert"
    assert assignment_role(_stage(trainees=[{"id": "a1"}]), "a1") == "Trainee Auditor"
    assert assignment_role(_stage(observers=[{"id": "a1"}]), "a1") == "Observer"
    assert assignment_role(
        _stage(lead_auditor_id="a1", technical_experts=[{"id": "a1"}]),
        "a1",
    ) == "Lead Auditor"


def test_calendar_rows_contain_only_assignments_and_group_consecutive_dates():
    occupied = _stage(technical_experts=[{"id": "a1"}])
    unrelated = _stage(
        audit_set=_audit_set(id="set-2", company_name="Other Company"),
        audit_date_start=date(2026, 9, 10),
        audit_date_end=date(2026, 9, 10),
        auditors=[{"id": "a2"}],
    )
    rows = calendar_rows([unrelated, occupied], "a1")

    assert len(rows) == 1
    assert rows[0]["audit_days"] == 3
    assert rows[0]["auditor_role"] == "Technical Expert"
    assert rows[0]["standards"] == ["ISO 9001:2015", "ISO 22000:2018"]
    assert rows[0]["audit_type"] == "Initial Certification"
    assert rows[0]["stage_type"] == "Stage 2"
    assert rows[0]["location"] == "Main Factory — 1 Factory Road, Istanbul"
    assert readable_date_range(rows[0]["date_start"], rows[0]["date_end"]) == "1–3 September 2026"


def test_surveillance_cycles_keep_their_exact_stage_number():
    assert stage_label("surveillance_1") == "Surveillance 1"
    assert stage_label("surveillance_2") == "Surveillance 2"
    assert stage_label("recertification") == "Recertification"


def test_excel_export_is_formatted_filterable_and_uses_typed_dates():
    rows = calendar_rows([_stage(lead_auditor_id="a1")], "a1")
    output = build_auditor_schedule_workbook("Aylin Auditor", rows)
    workbook = load_workbook(BytesIO(output))
    sheet = workbook["Audit Schedule"]

    assert sheet["A1"].value == "Auditor Schedule — Aylin Auditor"
    assert sheet["A5"].value == "1–3 September 2026"
    assert sheet["B5"].value == 3
    assert sheet["E5"].value == "ISO 9001:2015 + ISO 22000:2018"
    assert sheet["H5"].value == "Lead Auditor"
    assert sheet["J5"].value.date() == date(2026, 9, 1)
    assert sheet["K5"].value.date() == date(2026, 9, 3)
    assert sheet.column_dimensions["J"].hidden is True
    assert sheet.column_dimensions["K"].hidden is True
    assert sheet.freeze_panes == "A5"
    assert sheet.auto_filter.ref == "A4:K5"
    assert "AuditorSchedule" in sheet.tables


def test_excel_export_does_not_treat_user_text_as_a_formula():
    rows = calendar_rows([
        _stage(
            lead_auditor_id="a1",
            audit_set=_audit_set(company_name='=HYPERLINK("https://example.test","click")'),
        ),
    ], "a1")
    workbook = load_workbook(BytesIO(build_auditor_schedule_workbook("Auditor", rows)))
    company_cell = workbook["Audit Schedule"]["C5"]
    assert company_cell.data_type == "s"
    assert company_cell.value.startswith("'")
